import os
import io
import uuid
import datetime as dt

from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS

import pandas as pd
import joblib

from sqlalchemy import create_engine, Column, Integer, String, DateTime, Text, ForeignKey
from sqlalchemy.orm import declarative_base, sessionmaker, relationship
from sqlalchemy.exc import SQLAlchemyError

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# -----------------------------------------------------------
# FLASK
# -----------------------------------------------------------
app = Flask(__name__, static_folder="static", static_url_path="/static")
CORS(app, resources={r"/api/*": {"origins": "*"}})

TOOL_NAME = "TADILA"
TOOL_SIGLE = "TDL"

# -----------------------------------------------------------
# STORAGE / DB
# -----------------------------------------------------------
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///tadila.db")
engine = create_engine(DATABASE_URL, echo=False)
SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()

class FileRecord(Base):
    __tablename__ = "files"
    id = Column(Integer, primary_key=True)
    kind = Column(String(20))  # "training" | "processed"
    filename = Column(String(255))
    stored_name = Column(String(255))
    rows = Column(Integer, default=0)
    created_at = Column(DateTime, default=dt.datetime.utcnow)

    lines = relationship("DataRow", back_populates="file", cascade="all, delete-orphan")

class DataRow(Base):
    __tablename__ = "rows"
    id = Column(Integer, primary_key=True)
    file_id = Column(Integer, ForeignKey("files.id"))
    date = Column(String(100))
    sacs_id = Column(String(100))
    description = Column(Text)
    ro_label = Column(String(100))      # réel (training) ou attendu
    predicted_ro = Column(String(100))  # prediction (processed)
    kind = Column(String(20))           # "training" | "processed"

    file = relationship("FileRecord", back_populates="lines")

Base.metadata.create_all(engine)

# -----------------------------------------------------------
# MODELE (PKL / CSV)
# -----------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MODEL_PATH = os.path.join(BASE_DIR, "model.pkl")
VECT_PATH  = os.path.join(BASE_DIR, "vectorizer.pkl")
SEED_CSV   = os.path.join(BASE_DIR, "incidents.csv")

model = None
vectorizer = None

REQUIRED_COLS = ["Date", "SACS ID N°", "Description", "Règles d'or attribué"]

# -----------------------------------------------------------
# REGLES D'OR + ICONES
# Put PNGs in: backend/static/icons/RO1.png ...
# -----------------------------------------------------------
RULES = {
    "RO1":  {"fr":"Situations à haut risque","en":"High-risk situations","icon":"/static/icons/RO1.png"},
    "RO2":  {"fr":"Circulation","en":"Traffic","icon":"/static/icons/RO2.png"},
    "RO3":  {"fr":"Gestes & Outils","en":"Body mechanics & tools","icon":"/static/icons/RO3.png"},
    "RO4":  {"fr":"EPI","en":"PPE","icon":"/static/icons/RO4.png"},
    "RO5":  {"fr":"Permis de travail","en":"Work permits","icon":"/static/icons/RO5.png"},
    "RO6":  {"fr":"Opérations de levage","en":"Lifting operations","icon":"/static/icons/RO6.png"},
    "RO7":  {"fr":"Systèmes sous énergie","en":"Powered systems","icon":"/static/icons/RO7.png"},
    "RO8":  {"fr":"Espaces confinés","en":"Confined spaces","icon":"/static/icons/RO8.png"},
    "RO9":  {"fr":"Excavation","en":"Excavation","icon":"/static/icons/RO9.png"},
    "RO10": {"fr":"Travail en hauteur","en":"Work at height","icon":"/static/icons/RO10.png"},
    "RO11": {"fr":"Travaux par point chaud","en":"Hot work","icon":"/static/icons/RO11.png"},
    "RO12": {"fr":"Ligne de tir","en":"Line of fire","icon":"/static/icons/RO12.png"},
}

# -----------------------------------------------------------
# HELPERS
# -----------------------------------------------------------
def normalize_cols(df):
    df.columns = [str(c).strip() for c in df.columns]
    return df

def pick_col(df, candidates):
    norm_map = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        key = str(cand).strip().lower()
        if key in norm_map:
            return norm_map[key]
    return None

def find_required_cols(df):
    df = normalize_cols(df)
    date_col = pick_col(df, ["date", "Date"])
    sacs_col = pick_col(df, ["sacs id n°", "sacs id", "SACS ID N°"])
    desc_col = pick_col(df, ["description", "texte", "text", "Description"])
    ro_col   = pick_col(df, ["règles d'or attribué", "ro", "label", "Règles d'or attribué"])
    return date_col, sacs_col, desc_col, ro_col

def ensure_model_ready_or_error():
    if model is None or vectorizer is None or not hasattr(vectorizer, "vocabulary_"):
        raise RuntimeError("Le modèle n'est pas encore entraîné. Importez d'abord un fichier d'entraînement.")

def save_model(mdl, vect):
    global model, vectorizer
    model, vectorizer = mdl, vect
    joblib.dump(model, MODEL_PATH)
    joblib.dump(vectorizer, VECT_PATH)

def train_from_rows(texts, labels):
    vect = TfidfVectorizer()
    X = vect.fit_transform(texts)
    mdl = LogisticRegression(max_iter=2000).fit(X, labels)
    save_model(mdl, vect)

# -----------------------------------------------------------
# INIT MODEL
# -----------------------------------------------------------
def load_or_initialize_model():
    global model, vectorizer

    # 1) PKL
    if os.path.exists(MODEL_PATH) and os.path.exists(VECT_PATH):
        model = joblib.load(MODEL_PATH)
        vectorizer = joblib.load(VECT_PATH)
        return

    # 2) DB
    with SessionLocal() as s:
        rows = s.query(DataRow).filter(DataRow.kind == "training").all()
        if rows:
            texts = [r.description or "" for r in rows]
            labels = [r.ro_label or "" for r in rows]
            train_from_rows(texts, labels)
            return

    # 3) Seed CSV
    if os.path.exists(SEED_CSV):
        df = pd.read_csv(SEED_CSV)
        df = normalize_cols(df)
        _, _, desc_col, ro_col = find_required_cols(df)
        if desc_col and ro_col:
            texts = df[desc_col].astype(str).fillna("")
            labels = df[ro_col].astype(str).fillna("")
            train_from_rows(texts, labels)
            return

    # 4) Empty fallback
    vectorizer = TfidfVectorizer()
    model = LogisticRegression(max_iter=2000)

load_or_initialize_model()

# -----------------------------------------------------------
# FRONTEND SERVING
# -----------------------------------------------------------
@app.get("/")
def serve_ui():
    # serves backend/static/index.html
    return send_from_directory(app.static_folder, "index.html")

# optional: allow /index.html too
@app.get("/index.html")
def serve_ui_index():
    return send_from_directory(app.static_folder, "index.html")

# -----------------------------------------------------------
# API BASE
# -----------------------------------------------------------
@app.get("/api/health")
def api_health():
    return jsonify({"tool": TOOL_NAME, "sigle": TOOL_SIGLE, "status": "ok"})

@app.get("/api/rules")
def api_rules():
    return jsonify(RULES)

@app.get("/api/template")
def api_template():
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Modèle TADILA"
    ws.append(["Date", "SACS ID N°", "Description", "Règles d'or attribué"])
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="template_TADILA.xlsx", as_attachment=True)

@app.get("/api/stats")
def api_stats():
    with SessionLocal() as s:
        training_rows = s.query(DataRow).filter(DataRow.kind == "training").count()
        processed_rows = s.query(DataRow).filter(DataRow.kind == "processed").count()
    return jsonify({"training_rows": training_rows, "processed_rows": processed_rows})

@app.get("/api/files")
def api_files():
    kind = request.args.get("kind", "").strip().lower()
    if kind not in ("training", "processed"):
        return jsonify({"error": "Query param 'kind' must be 'training' or 'processed'"}), 400

    with SessionLocal() as s:
        files = (
            s.query(FileRecord)
            .filter(FileRecord.kind == kind)
            .order_by(FileRecord.created_at.desc())
            .all()
        )

    out = []
    for f in files:
        out.append({
            "id": f.id,
            "kind": f.kind,
            "filename": f.filename,
            "rows": f.rows,
            "created_at": f.created_at.isoformat() + "Z" if f.created_at else None
        })

    return jsonify({"status": "ok", "files": out})

# -----------------------------------------------------------
# UPLOAD TRAINING
# -----------------------------------------------------------
@app.post("/api/upload-training")
def api_upload_training():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Aucun fichier fourni"}), 400

    try:
        df = pd.read_excel(file)
    except Exception as e:
        return jsonify({"error": f"Lecture Excel impossible: {e}"}), 400

    df = normalize_cols(df)
    date_col, sacs_col, desc_col, ro_col = find_required_cols(df)
    if not (date_col and sacs_col and desc_col and ro_col):
        return jsonify({
            "error": "Colonnes requises manquantes",
            "obligatoires": REQUIRED_COLS,
            "trouvees": list(df.columns)
        }), 400

    inserted = 0
    try:
        with SessionLocal() as s:
            f = FileRecord(kind="training", filename=getattr(file, "filename", "upload.xlsx"),
                           stored_name=f"training_{uuid.uuid4().hex}.xlsx")
            s.add(f)
            s.flush()

            for _, row in df.iterrows():
                s.add(DataRow(
                    file_id=f.id,
                    date=str(row.get(date_col, "")),
                    sacs_id=str(row.get(sacs_col, "")),
                    description=str(row.get(desc_col, "")),
                    ro_label=str(row.get(ro_col, "")),
                    kind="training"
                ))
                inserted += 1

            f.rows = inserted
            s.commit()

        # retrain from all training rows
        with SessionLocal() as s:
            rows = s.query(DataRow).filter(DataRow.kind == "training").all()
            texts  = [r.description or "" for r in rows]
            labels = [r.ro_label or "" for r in rows]
        train_from_rows(texts, labels)

    except SQLAlchemyError as e:
        return jsonify({"error": f"Insertion DB échouée: {e}"}), 500
    except Exception as e:
        return jsonify({"error": f"Réentraînement échoué: {e}"}), 500

    return jsonify({"status": "ok", "message": "Importation successfully", "inserted_rows": inserted})

# -----------------------------------------------------------
# PROCESS DATA
# -----------------------------------------------------------
@app.post("/api/process-data")
def api_process_data():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Aucun fichier fourni"}), 400

    try:
        df = pd.read_excel(file)
    except Exception as e:
        return jsonify({"error": f"Lecture Excel impossible: {e}"}), 400

    df = normalize_cols(df)
    date_col, sacs_col, desc_col, ro_col = find_required_cols(df)
    if not (date_col and sacs_col and desc_col):
        return jsonify({
            "error": "Colonnes requises manquantes",
            "obligatoires_min": ["Date", "SACS ID N°", "Description"],
            "trouvees": list(df.columns)
        }), 400

    try:
        ensure_model_ready_or_error()
        X = vectorizer.transform(df[desc_col].astype(str).fillna(""))
        preds = model.predict(X)
    except Exception as e:
        return jsonify({"error": f"Analyse impossible: {e}"}), 500

    inserted = 0
    file_id = None

    try:
        with SessionLocal() as s:
            f = FileRecord(kind="processed", filename=getattr(file, "filename", "source.xlsx"),
                           stored_name=f"processed_{uuid.uuid4().hex}.xlsx")
            s.add(f)
            s.flush()
            file_id = f.id

            for i, row in df.iterrows():
                s.add(DataRow(
                    file_id=f.id,
                    date=str(row.get(date_col, "")),
                    sacs_id=str(row.get(sacs_col, "")),
                    description=str(row.get(desc_col, "")),
                    ro_label=str(row.get(ro_col, "")) if ro_col in df.columns else "",
                    predicted_ro=str(preds[i]),
                    kind="processed"
                ))
                inserted += 1

            f.rows = inserted
            s.commit()

    except SQLAlchemyError as e:
        return jsonify({"error": f"Sauvegarde DB échouée: {e}"}), 500

    # Build preview (first 50 rows)
    preview = []
    for i in range(min(50, len(df))):
        ro = str(preds[i]).strip()
        icon = RULES.get(ro, {}).get("icon", "")
        preview.append({
            "Date": str(df.iloc[i].get(date_col, "")),
            "SACS ID N°": str(df.iloc[i].get(sacs_col, "")),
            "Description": str(df.iloc[i].get(desc_col, "")),
            "RO_predite": ro,
            "RO_icon": icon
        })

    return jsonify({
        "status": "ok",
        "file_id": file_id,
        "job_id": file_id,              # alias for compatibility
        "rows": inserted,               # what frontend expects
        "inserted_rows": inserted,      # also keep original
        "preview": preview
    })

# -----------------------------------------------------------
# EXPORT EXCEL
# -----------------------------------------------------------
@app.get("/api/export/excel/<int:file_id>")
def api_export_excel(file_id):
    with SessionLocal() as s:
        f = s.query(FileRecord).filter(FileRecord.id == file_id, FileRecord.kind == "processed").first()
        if not f:
            return jsonify({"error": "Fichier traité introuvable"}), 404
        rows = s.query(DataRow).filter(DataRow.file_id == file_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Analysé"

    headers = ["Date", "SACS ID N°", "Description", "RO (attendu)", "RO (prédit)", "Icône"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="009543")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    r = 2
    for line in rows:
        ws.append([
            line.date or "",
            line.sacs_id or "",
            line.description or "",
            line.ro_label or "",
            line.predicted_ro or "",
            ""
        ])

        ro = (line.predicted_ro or "").strip()
        icon_rel = RULES.get(ro, {}).get("icon")
        if icon_rel:
            # icon_rel: /static/icons/RO1.png
            icon_abs = os.path.join(BASE_DIR, icon_rel.lstrip("/"))
            if os.path.exists(icon_abs):
                img = XLImage(icon_abs)
                img.width, img.height = 20, 20
                ws.add_image(img, f"F{r}")
        r += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name=f"tadila_export_{file_id}.xlsx", as_attachment=True)

# -----------------------------------------------------------
# RUN
# -----------------------------------------------------------
if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=True)


